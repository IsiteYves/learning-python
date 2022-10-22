import openpyxl as xl
from openpyxl.chart import BarChart, Reference

XLfile=xl.load_workbook('salaries.xlsx')

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    correct_val = cell.value * 0.814
    sheet.cell(row, 4).value = correct_val
    
content = Reference(sheet, min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart=BarChart()
chart.add_data(content)
sheet.add_chart(chart, 'f5')
    
XLfile.save('salaries.xlsx')



